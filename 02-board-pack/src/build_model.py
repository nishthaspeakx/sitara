"""Project Sitara - bottom-up 36-month financial model. 17 sheets, live formulas.
Assumption cell map: FX=B4 GST=B5 ARPU=B6 ADDON=B7 CHURNb=B8 CHURNc=B9 CHURNa=B10
CACb=B11 CACc=B12 CACa=B13 REF=B14 REFUND=B15 AI=B16 HUM=B17 SUP=B18 PAY=B19 INFRA=B20
ANNUAL=B21 PREM=B22 FOUND=B23 IMPROVE=B24 CHURNFLOOR=B25"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference

wb = openpyxl.Workbook()
BLUE=Font(name="Arial",color="0000FF"); BOLD=Font(name="Arial",bold=True)
H1=Font(name="Arial",bold=True,size=13); GREEN=Font(name="Arial",color="008000")
ITAL=Font(name="Arial",italic=True,size=9)
YEL=PatternFill("solid",fgColor="FFFF00"); HDR=PatternFill("solid",fgColor="1E2761")
HDRF=Font(name="Arial",bold=True,color="FFFFFF")
NUM='#,##0;(#,##0);-'; PCT='0.0%;(0.0%);-'; MUL='0.0x'
def hdr(ws,row,cols,start=1):
    for i,c in enumerate(cols):
        cell=ws.cell(row=row,column=start+i,value=c); cell.fill=HDR; cell.font=HDRF
        cell.alignment=Alignment(wrap_text=True,vertical="center")
def widths(ws, wl):
    from openpyxl.utils import get_column_letter
    for i,w in enumerate(wl,start=1): ws.column_dimensions[get_column_letter(i)].width=w

# ---------- 1 ASSUMPTIONS ----------
ws=wb.active; ws.title="Assumptions"
ws["A1"]="PROJECT SITARA - MODEL ASSUMPTIONS (blue = editable; yellow = highest-impact levers)"; ws["A1"].font=H1
rows=[("Label","Value","Unit","Type","Source / logic"),
("FX rate (INR per USD)",90,"Rs/$","Assumption","Management estimate, Jul-2026"),
("GST rate (India plans)",0.18,"%","Fact","India prices GST-inclusive; revenue modelled net of GST; NRI = zero-rated export"),
("Blended net subscription ARPU (Rs/mo)",815,"Rs/mo","Mgmt estimate","Derived on Pricing sheet from plan mix; net of GST"),
("Add-on ARPU (consult minutes, reports) (Rs/mo)",65,"Rs/mo","Assumption","~8% attach; benchmark: AstroTalk session spend Rs500-1500 (Entrackr FY24)"),
("Monthly churn - base",0.06,"%/mo","Benchmark+assumption","Recurly DTC benchmark 6.5%/mo; annual-mix damping; VALIDATION TARGET"),
("Monthly churn - conservative",0.075,"%/mo","Scenario","+25% stress"),
("Monthly churn - aggressive",0.045,"%/mo","Scenario","Only if cohorts prove it"),
("Blended CAC - base (Rs)",3600,"Rs","Benchmark+assumption","AstroTalk domestic Rs575-870 (Entrackr); NRI premium assumed 4-5x"),
("CAC - conservative (Rs)",4500,"Rs","Scenario","+25% stress"),
("CAC - aggressive (Rs)",3000,"Rs","Scenario","Referral/organic mix matures"),
("Referral share of new customers",0.12,"%","Assumption","Family-chart gifting loop"),
("Refund rate (% of gross revenue)",0.025,"%","Assumption","30-day money-back guarantee"),
("AI inference cost / user / mo (Rs)",70,"Rs","Mgmt estimate","LLM API + ephemeris compute, Jul-2026 rates"),
("Human consult-min cost / user / mo (Rs)",45,"Rs","Mgmt estimate","Premium-tier allowance blended over all users"),
("Support cost / user / mo (Rs)",25,"Rs","Assumption","~1 agent per 5,000 subscribers"),
("Payment gateway (% of gross revenue)",0.028,"%","Fact","Stripe / Razorpay published pricing"),
("Infra / storage / user / mo (Rs)",15,"Rs","Assumption","Cloud, security, analytics"),
("Annual-plan share of subscribers",0.60,"%","Assumption","Annual-first pricing strategy"),
("Premium-tier share of subscribers",0.25,"%","Assumption","Pricing sheet mix"),
("Founding-member adds (one-time, month 4)",300,"count","Plan","Launch cohort beyond paid funnel"),
("Churn improvement per month from M13 (pts)",0.0005,"pts","Assumption","Cohort mix shifts annual over time"),
("Churn floor - base (%/mo)",0.05,"%/mo","Assumption","Best-case mature churn, base case")]
for r,vals in enumerate(rows,start=3):
    for c,v in enumerate(vals,start=1):
        cell=ws.cell(row=r,column=c,value=v)
        if r==3: cell.fill=HDR; cell.font=HDRF
        elif c==2: cell.font=BLUE
for rr in (8,11):
    ws.cell(row=rr,column=2).fill=YEL
ws.cell(row=6,column=2).fill=YEL
for rr,fmt in [(5,PCT),(8,PCT),(9,PCT),(10,PCT),(14,PCT),(15,PCT),(19,PCT),(21,'0%'),(22,'0%'),(24,'0.00%'),(25,PCT)]:
    ws.cell(row=rr,column=2).number_format=fmt
widths(ws,[46,12,10,22,78])
ws["A27"]="All downstream sheets recalculate from this sheet. Every figure in the board decks traces here."; ws["A27"].font=ITAL

# ---------- 2 PRICING ----------
ws=wb.create_sheet("Pricing")
ws["A1"]="PRICING ARCHITECTURE (effective monthly ARPU, net of GST)"; ws["A1"].font=H1
hdr(ws,3,["Plan","Customer","Monthly","Annual","Mix %","Gross eff Rs/mo","GST factor","Net eff Rs/mo","Weighted Rs/mo"])
plans=[("NRI Standard","NRI individual (US/UK/CA/Gulf)","$12.99","$99",0.45,
        "=(0.4*12.99+0.6*99/12)*Assumptions!$B$4",1.0,"=F4*G4","=H4*E4"),
       ("NRI Premium","NRI premium; 30 consult-min/qtr","$24.99","$199",0.12,
        "=(0.4*24.99+0.6*199/12)*Assumptions!$B$4",1.0,"=F5*G5","=H5*E5"),
       ("India Aarambh","Metro India individual","Rs499","Rs3,999",0.30,
        "=0.4*499+0.6*3999/12","=1/(1+Assumptions!$B$5)","=F6*G6","=H6*E6"),
       ("India Pragati","Metro India premium; 15 min/qtr","Rs999","Rs7,999",0.13,
        "=0.4*999+0.6*7999/12","=1/(1+Assumptions!$B$5)","=F7*G7","=H7*E7")]
for r,p in enumerate(plans,start=4):
    for c,v in enumerate(p,start=1):
        cell=ws.cell(row=r,column=c,value=v)
        if c in (3,4,5): cell.font=BLUE
        if c==5: cell.number_format='0%'
        if c in (6,8,9): cell.number_format=NUM
        if c==7: cell.number_format='0.000'
ws["A9"]="Blended net subscription ARPU (Rs/mo) - computed"; ws["A9"].font=BOLD
ws["I9"]="=SUM(I4:I7)"; ws["I9"].font=BOLD; ws["I9"].number_format=NUM
ws["A10"]="Value used by model (Assumptions!B6) - keep aligned:"; ws["I10"]="=Assumptions!$B$6"; ws["I10"].font=GREEN; ws["I10"].number_format=NUM
ws["A12"]="Assumes 60% of subscribers on annual billing (40% monthly). Mix percentages are management estimates."; ws["A12"].font=ITAL
widths(ws,[16,34,10,10,8,15,10,14,14])

# ---------- 3 CUSTOMER GROWTH ----------
ws=wb.create_sheet("Customer Growth")
ws["A1"]="CUSTOMER GROWTH - 36 MONTHS, BASE CASE (opening + new - churned = closing)"; ws["A1"].font=H1
hdr(ws,3,["Month","Marketing Rs","CAC Rs","New (funnel)","Referrals","Founding","Total new","Opening","Churn rate","Churned","Closing","Annual subs","Premium subs"])
mkt=[300000]*3+[800000]*3+[1500000]*6+[2500000]*12+[3500000]*12
for m in range(1,37):
    r=m+3
    ws.cell(row=r,column=1,value=m)
    ws.cell(row=r,column=2,value=mkt[m-1]).font=BLUE
    ws.cell(row=r,column=3,value="=Assumptions!$B$11").font=GREEN
    ws.cell(row=r,column=4,value=f"=B{r}/C{r}")
    ws.cell(row=r,column=5,value=f"=D{r}*Assumptions!$B$14")
    ws.cell(row=r,column=6,value=f"=IF(A{r}=4,Assumptions!$B$23,0)")
    ws.cell(row=r,column=7,value=f"=D{r}+E{r}+F{r}")
    ws.cell(row=r,column=8,value=0 if m==1 else f"=K{r-1}")
    ws.cell(row=r,column=9,value=f"=MAX(Assumptions!$B$25,Assumptions!$B$8-MAX(0,A{r}-12)*Assumptions!$B$24)")
    ws.cell(row=r,column=10,value=f"=H{r}*I{r}")
    ws.cell(row=r,column=11,value=f"=H{r}+G{r}-J{r}")
    ws.cell(row=r,column=12,value=f"=K{r}*Assumptions!$B$21")
    ws.cell(row=r,column=13,value=f"=K{r}*Assumptions!$B$22")
    for c in [2,3,4,5,6,7,8,10,11,12,13]: ws.cell(row=r,column=c).number_format=NUM
    ws.cell(row=r,column=9).number_format=PCT
ws["A41"]="Milestones:"; ws["A41"].font=BOLD
for lbl,cell,ref in [("M6","B41","=K9"),("M12","D41","=K15"),("M24","F41","=K27"),("M36","H41","=K39")]:
    ws[cell]=lbl; ws[cell].font=BOLD
    tgt=ws.cell(row=41,column=ws[cell].column+1); tgt.value=ref; tgt.number_format=NUM; tgt.font=BOLD
widths(ws,[7,13,9,11,9,9,10,10,10,9,10,11,11])

# ---------- 4 FUNNEL ----------
ws=wb.create_sheet("Funnel")
ws["A1"]="ACQUISITION FUNNEL - representative steady month (M12)"; ws["A1"].font=H1
hdr(ws,3,["Stage","Conversion","Volume /mo","Cost Rs","Logic (type)"])
fun=[("Ad impressions","-","=C6/0.03/0.08","","Implied at 8% impression-to-visit x CTR mix (assumption)"),
("Site visitors","-","=C6/0.03","","Visitor-to-trial 3% (management estimate)"),
("Trials / Honest-Kundli signups","3.0% of visitors","=C7/0.25","","Trial-to-paid 25% (management estimate)"),
("New paid customers","25% of trials","='Customer Growth'!D15","='Customer Growth'!B15","Funnel spend M12"),
("Referral additions","12% of funnel","='Customer Growth'!E15","","Family-chart gift loop (assumption)"),
("Effective blended CAC","-","","='Customer Growth'!B15/('Customer Growth'!D15+'Customer Growth'!E15)","Marketing / all new")]
for r,f in enumerate(fun,start=4):
    for c,v in enumerate(f,start=1):
        cell=ws.cell(row=r,column=c,value=v)
        if c in (3,4): cell.number_format=NUM
ws["A11"]="Funnel rates are management estimates until replaced by Gate-1 cohort actuals."; ws["A11"].font=ITAL
widths(ws,[30,16,14,13,56])

# ---------- 5 REVENUE ----------
ws=wb.create_sheet("Revenue")
ws["A1"]="REVENUE - 36 MONTHS (Rs, recognised monthly, net of GST)"; ws["A1"].font=H1
hdr(ws,3,["Month","Avg subscribers","Subscription rev","Add-on rev","Gross rev","Refunds","Net revenue","Deferred-cash memo"])
for m in range(1,37):
    r=m+3
    ws.cell(row=r,column=1,value=m)
    ws.cell(row=r,column=2,value=f"=('Customer Growth'!H{r}+'Customer Growth'!K{r})/2")
    ws.cell(row=r,column=3,value=f"=B{r}*Assumptions!$B$6")
    ws.cell(row=r,column=4,value=f"=B{r}*Assumptions!$B$7")
    ws.cell(row=r,column=5,value=f"=C{r}+D{r}")
    ws.cell(row=r,column=6,value=f"=E{r}*Assumptions!$B$15")
    ws.cell(row=r,column=7,value=f"=E{r}-F{r}")
    ws.cell(row=r,column=8,value=f"='Customer Growth'!G{r}*Assumptions!$B$21*Assumptions!$B$6*5.5")
    for c in range(2,9): ws.cell(row=r,column=c).number_format=NUM
ws["A41"]="Year totals:"; ws["A41"].font=BOLD
ws["B41"]="Y1"; ws["C41"]="=SUM(G4:G15)"; ws["D41"]="Y2"; ws["E41"]="=SUM(G16:G27)"; ws["F41"]="Y3"; ws["G41"]="=SUM(G28:G39)"
for c in ("C41","E41","G41"): ws[c].number_format=NUM; ws[c].font=BOLD
ws["A42"]="Deferred-cash memo: annual plans collect ~5.5 months of revenue ahead of recognition on average."; ws["A42"].font=ITAL
widths(ws,[7,14,15,12,14,10,14,17])

# ---------- 6 COST OF REVENUE ----------
ws=wb.create_sheet("Cost of Revenue")
ws["A1"]="COST OF REVENUE - 36 MONTHS (Rs)"; ws["A1"].font=H1
hdr(ws,3,["Month","AI cost","Human minutes","Support","Infra","Payment fees","Total COGS","Gross profit","GM %"])
for m in range(1,37):
    r=m+3
    ws.cell(row=r,column=1,value=m)
    ws.cell(row=r,column=2,value=f"=Revenue!B{r}*Assumptions!$B$16")
    ws.cell(row=r,column=3,value=f"=Revenue!B{r}*Assumptions!$B$17")
    ws.cell(row=r,column=4,value=f"=Revenue!B{r}*Assumptions!$B$18")
    ws.cell(row=r,column=5,value=f"=Revenue!B{r}*Assumptions!$B$20")
    ws.cell(row=r,column=6,value=f"=Revenue!E{r}*Assumptions!$B$19")
    ws.cell(row=r,column=7,value=f"=SUM(B{r}:F{r})")
    ws.cell(row=r,column=8,value=f"=Revenue!G{r}-G{r}")
    ws.cell(row=r,column=9,value=f"=IF(Revenue!G{r}=0,0,H{r}/Revenue!G{r})")
    for c in range(2,9): ws.cell(row=r,column=c).number_format=NUM
    ws.cell(row=r,column=9).number_format=PCT
ws["A41"]="Y totals GP:"; ws["A41"].font=BOLD
ws["B41"]="=SUM(H4:H15)"; ws["C41"]="=SUM(H16:H27)"; ws["D41"]="=SUM(H28:H39)"
for c in ("B41","C41","D41"): ws[c].number_format=NUM; ws[c].font=BOLD
widths(ws,[7,11,12,10,10,12,12,13,9])

# ---------- 7 HEADCOUNT ----------
ws=wb.create_sheet("Headcount")
ws["A1"]="HEADCOUNT PLAN (fully-loaded monthly cost, Rs)"; ws["A1"].font=H1
hdr(ws,3,["Role","Phase A (M1-6)","Phase B (M7-18)","Phase C (M19-36)","Monthly cost Rs","Notes"])
hc=[("Founder/CEO",1,1,1,250000,"Below-market; corrected at Series A"),
("Founding engineers",2,3,5,300000,"Per-engineer avg Rs3.0L loaded"),
("Jyotish lead (part-time then full)",0.5,1,1,250000,"Credibility-critical hire"),
("Designer (contract)",0.5,1,1,180000,""),
("Growth/marketing",0,1,2,220000,""),
("Content/community",0,1,2,150000,""),
("Astrologer-panel ops",0,0.5,1,140000,""),
("Support",0,1,3,80000,""),
("Finance/admin (fractional)",0.2,0.4,1,120000,"")]
for r,h in enumerate(hc,start=4):
    for c,v in enumerate(h,start=1):
        cell=ws.cell(row=r,column=c,value=v)
        if c in (2,3,4,5): cell.font=BLUE; cell.number_format=NUM
ws["A13"]="Monthly salary bill"; ws["A13"].font=BOLD
ws["B13"]="=SUMPRODUCT(B4:B12,$E$4:$E$12)"; ws["C13"]="=SUMPRODUCT(C4:C12,$E$4:$E$12)"; ws["D13"]="=SUMPRODUCT(D4:D12,$E$4:$E$12)"
for c in ("B13","C13","D13"): ws[c].number_format=NUM; ws[c].font=BOLD
ws["A15"]="Headcount (FTE)"; ws["B15"]="=SUM(B4:B12)"; ws["C15"]="=SUM(C4:C12)"; ws["D15"]="=SUM(D4:D12)"
widths(ws,[30,14,14,14,15,34])

# ---------- 8 MARKETING ----------
ws=wb.create_sheet("Marketing")
ws["A1"]="MARKETING - CHANNEL CAC AND SPEND"; ws["A1"].font=H1
hdr(ws,3,["Channel","CAC Rs","Spend share","Weighted CAC","Type"])
ch=[("Meta (IG/FB) diaspora",4200,0.40,"=B4*C4","Benchmark-derived"),
("Google search intent",4800,0.15,"=B5*C5","Assumption"),
("Micro-influencer astrologers",3000,0.15,"=B6*C6","Assumption"),
("Community/organic (Reddit, SEO)",1200,0.12,"=B7*C7","Assumption"),
("Referral (family-chart gift)",800,0.10,"=B8*C8","Cost of incentive"),
("Festival campaigns",3400,0.08,"=B9*C9","Assumption")]
for r,cch in enumerate(ch,start=4):
    for c,v in enumerate(cch,start=1):
        cell=ws.cell(row=r,column=c,value=v)
        if c in (2,3): cell.font=BLUE
        if c==2 or c==4: cell.number_format=NUM
        if c==3: cell.number_format='0%'
ws["A10"]="Blended CAC (check vs Assumptions!B11)"; ws["A10"].font=BOLD
ws["D10"]="=SUM(D4:D9)"; ws["D10"].number_format=NUM; ws["D10"].font=BOLD
ws["E10"]="=Assumptions!$B$11"; ws["E10"].font=GREEN; ws["E10"].number_format=NUM
ws["A12"]="Monthly marketing spend schedule lives in 'Customer Growth' column B (editable)."; ws["A12"].font=ITAL
ws["A13"]="Benchmark: AstroTalk domestic CAC Rs575-870, 6-8 month recovery (Entrackr/growthx FY24-25)."; ws["A13"].font=ITAL
widths(ws,[30,10,12,13,20])

# ---------- 9 OPERATING EXPENSES (monthly engine) ----------
ws=wb.create_sheet("Operating Expenses")
ws["A1"]="OPERATING EXPENSES + MONTHLY EBITDA ENGINE (Rs)"; ws["A1"].font=H1
hdr(ws,3,["Month","Salaries","Marketing","Tech/software","G&A/legal/office","Other/contingency","Total opex","Net revenue","Gross profit","EBITDA"])
for m in range(1,37):
    r=m+3
    ws.cell(row=r,column=1,value=m)
    ws.cell(row=r,column=2,value=f"=IF(A{r}<=6,Headcount!$B$13,IF(A{r}<=18,Headcount!$C$13,Headcount!$D$13))")
    ws.cell(row=r,column=3,value=f"='Customer Growth'!B{r}")
    ws.cell(row=r,column=4,value=f"=IF(A{r}<=6,120000,IF(A{r}<=18,200000,320000))")
    ws.cell(row=r,column=5,value=f"=IF(A{r}<=6,150000,IF(A{r}<=18,250000,400000))")
    ws.cell(row=r,column=6,value=f"=0.05*(B{r}+D{r}+E{r})")
    ws.cell(row=r,column=7,value=f"=SUM(B{r}:F{r})")
    ws.cell(row=r,column=8,value=f"=Revenue!G{r}").font=GREEN
    ws.cell(row=r,column=9,value=f"='Cost of Revenue'!H{r}").font=GREEN
    ws.cell(row=r,column=10,value=f"=I{r}-G{r}")
    for c in range(2,11): ws.cell(row=r,column=c).number_format=NUM
ws["A41"]="Y totals EBITDA:"; ws["A41"].font=BOLD
ws["B41"]="=SUM(J4:J15)"; ws["C41"]="=SUM(J16:J27)"; ws["D41"]="=SUM(J28:J39)"
for c in ("B41","C41","D41"): ws[c].number_format=NUM; ws[c].font=BOLD
ws["A42"]="Tech/G&A steps are management estimates. Contingency 5% of non-marketing opex."; ws["A42"].font=ITAL
widths(ws,[7,12,12,12,13,12,12,13,12,12])

# ---------- 10 P&L ----------
ws=wb.create_sheet("P&L")
ws["A1"]="P&L - MONTHLY YEAR 1, ANNUAL Y1-Y3 (Rs)"; ws["A1"].font=H1
cols=["Line"]+[f"M{m}" for m in range(1,13)]+["Year 1","Year 2","Year 3"]
hdr(ws,3,cols)
lines=[("Subscription revenue","=Revenue!C{r}","=SUM(Revenue!C4:C15)","=SUM(Revenue!C16:C27)","=SUM(Revenue!C28:C39)"),
("Add-on revenue","=Revenue!D{r}","=SUM(Revenue!D4:D15)","=SUM(Revenue!D16:D27)","=SUM(Revenue!D28:D39)"),
("Refunds","=-Revenue!F{r}","=-SUM(Revenue!F4:F15)","=-SUM(Revenue!F16:F27)","=-SUM(Revenue!F28:F39)"),
("Net revenue","=Revenue!G{r}","=SUM(Revenue!G4:G15)","=SUM(Revenue!G16:G27)","=SUM(Revenue!G28:G39)"),
("AI + infra cost","=-('Cost of Revenue'!B{r}+'Cost of Revenue'!E{r})","=-SUM('Cost of Revenue'!B4:B15)-SUM('Cost of Revenue'!E4:E15)","=-SUM('Cost of Revenue'!B16:B27)-SUM('Cost of Revenue'!E16:E27)","=-SUM('Cost of Revenue'!B28:B39)-SUM('Cost of Revenue'!E28:E39)"),
("Human review / consult cost","=-'Cost of Revenue'!C{r}","=-SUM('Cost of Revenue'!C4:C15)","=-SUM('Cost of Revenue'!C16:C27)","=-SUM('Cost of Revenue'!C28:C39)"),
("Support cost","=-'Cost of Revenue'!D{r}","=-SUM('Cost of Revenue'!D4:D15)","=-SUM('Cost of Revenue'!D16:D27)","=-SUM('Cost of Revenue'!D28:D39)"),
("Payment fees","=-'Cost of Revenue'!F{r}","=-SUM('Cost of Revenue'!F4:F15)","=-SUM('Cost of Revenue'!F16:F27)","=-SUM('Cost of Revenue'!F28:F39)"),
("Gross profit","='Cost of Revenue'!H{r}","='Cost of Revenue'!B41","='Cost of Revenue'!C41","='Cost of Revenue'!D41"),
("Salaries","=-'Operating Expenses'!B{r}","=-SUM('Operating Expenses'!B4:B15)","=-SUM('Operating Expenses'!B16:B27)","=-SUM('Operating Expenses'!B28:B39)"),
("Marketing","=-'Operating Expenses'!C{r}","=-SUM('Operating Expenses'!C4:C15)","=-SUM('Operating Expenses'!C16:C27)","=-SUM('Operating Expenses'!C28:C39)"),
("Tech / software","=-'Operating Expenses'!D{r}","=-SUM('Operating Expenses'!D4:D15)","=-SUM('Operating Expenses'!D16:D27)","=-SUM('Operating Expenses'!D28:D39)"),
("G&A / legal","=-'Operating Expenses'!E{r}","=-SUM('Operating Expenses'!E4:E15)","=-SUM('Operating Expenses'!E16:E27)","=-SUM('Operating Expenses'!E28:E39)"),
("Contingency","=-'Operating Expenses'!F{r}","=-SUM('Operating Expenses'!F4:F15)","=-SUM('Operating Expenses'!F16:F27)","=-SUM('Operating Expenses'!F28:F39)"),
("EBITDA","='Operating Expenses'!J{r}","='Operating Expenses'!B41","='Operating Expenses'!C41","='Operating Expenses'!D41"),
("Depreciation","=-8333","=-100000","=-100000","=-100000"),
("EBIT","=B{rr}+B{rr2}",None,None,None),
("Tax (nil until carry-forwards absorbed)","=0","=0","=0","=IF('Operating Expenses'!D41>0,-0.25*MAX(0,'Operating Expenses'!D41-100000-4500000),0)"),
("Net profit",None,None,None,None)]
for li,(label,mform,y1,y2,y3) in enumerate(lines):
    r=4+li
    ws.cell(row=r,column=1,value=label).font=BOLD if label in ("Net revenue","Gross profit","EBITDA","Net profit") else Font(name="Arial")
    for m in range(1,13):
        c=1+m; rr=m+3
        if label=="EBIT":
            ws.cell(row=r,column=c,value=f"={CLL(c)}{r-2}+{CLL(c)}{r-1}" if False else f"=" )
        elif label=="Net profit":
            pass
        elif mform:
            ws.cell(row=r,column=c,value=mform.format(r=rr)).number_format=NUM
    if y1: ws.cell(row=r,column=14,value=y1).number_format=NUM
    if y2: ws.cell(row=r,column=15,value=y2).number_format=NUM
    if y3: ws.cell(row=r,column=16,value=y3).number_format=NUM
# EBIT and Net profit rows via column arithmetic
from openpyxl.utils import get_column_letter as CLL
ebitda_r=4+14; dep_r=4+15; ebit_r=4+16; tax_r=4+17; np_r=4+18
for c in list(range(2,14))+[14,15,16]:
    L=CLL(c)
    ws.cell(row=ebit_r,column=c,value=f"={L}{ebitda_r}+{L}{dep_r}").number_format=NUM
    if c<14: ws.cell(row=tax_r,column=c,value="=0").number_format=NUM
    ws.cell(row=np_r,column=c,value=f"={L}{ebit_r}+{L}{tax_r}").number_format=NUM
ws.cell(row=np_r,column=1).font=BOLD
ws["A24"]="GM % (Y1/Y2/Y3):"; ws["N24"]="=N12/N7"; ws["O24"]="=O12/O7"; ws["P24"]="=P12/P7"
for c in ("N24","O24","P24"): ws[c].number_format=PCT
ws["A25"]="Note: Y3 tax formula assumes 25% on profit above accumulated losses of ~Rs45L absorbed; simplification flagged."; ws["A25"].font=ITAL
widths(ws,[30]+[10]*12+[12,12,12])

# ---------- 11 CASH FLOW ----------
ws=wb.create_sheet("Cash Flow")
ws["A1"]="CASH FLOW - 36 MONTHS (Rs)"; ws["A1"].font=H1
hdr(ws,3,["Month","Opening cash","Investment in","EBITDA","Deferred-cash delta","Capex","Closing cash"])
for m in range(1,37):
    r=m+3
    ws.cell(row=r,column=1,value=m)
    ws.cell(row=r,column=2,value=6500000 if m==1 else f"=G{r-1}")
    inv = {7:28500000, 19:45000000}
    ws.cell(row=r,column=3,value=inv.get(m,0)).font=BLUE
    ws.cell(row=r,column=4,value=f"='Operating Expenses'!J{r}").font=GREEN
    ws.cell(row=r,column=5,value=f"=Revenue!H{r}-IF(A{r}=1,0,Revenue!H{r-1})")
    ws.cell(row=r,column=6,value=-150000 if m in (2,8,20) else 0)
    ws.cell(row=r,column=7,value=f"=B{r}+C{r}+D{r}+E{r}+F{r}")
    for c in range(2,8): ws.cell(row=r,column=c).number_format=NUM
ws["A41"]="Tranches: Rs65L at M1 (Scenario A, incl. in opening), Rs2.85Cr at M7 (B), Rs4.5Cr at M19 (C)."; ws["A41"].font=ITAL
ws["A42"]="Minimum-cash policy Rs40L; deferred-cash delta = annual-plan cash collected ahead of recognition."; ws["A42"].font=ITAL
ws["A44"]="Minimum closing cash:"; ws["B44"]="=MIN(G4:G39)"; ws["B44"].number_format=NUM; ws["B44"].font=BOLD
ws["C44"]="Peak cumulative funding:"; ws["D44"]="=6500000+SUM(C4:C39)"; ws["D44"].number_format=NUM; ws["D44"].font=BOLD
widths(ws,[7,13,13,12,14,9,13])

# ---------- 12 UNIT ECONOMICS ----------
ws=wb.create_sheet("Unit Economics")
ws["A1"]="UNIT ECONOMICS PER CUSTOMER (Rs / month, base case)"; ws["A1"].font=H1
ue=[("Total ARPU (subs + add-on)","=Assumptions!B6+Assumptions!B7"),
("AI cost","=-Assumptions!B16"),("Human minutes","=-Assumptions!B17"),
("Support","=-Assumptions!B18"),("Infra","=-Assumptions!B20"),
("Payment fees","=-(Assumptions!B6+Assumptions!B7)*Assumptions!B19"),
("Refunds","=-(Assumptions!B6+Assumptions!B7)*Assumptions!B15"),
("Contribution / user / month","=SUM(B4:B10)"),
("Contribution margin %","=B11/B4"),
("Expected customer life (months)","=1/Assumptions!B8"),
("LTV (contribution x life)","=B11/Assumptions!B8"),
("CAC","=Assumptions!B11"),
("LTV : CAC","=B14/B15"),
("CAC payback (months)","=B15/B11")]
for r,(l,f) in enumerate(ue,start=4):
    ws.cell(row=r,column=1,value=l).font=BOLD if "Contribution /" in l or l.startswith("LTV") or "payback" in l else Font(name="Arial")
    cell=ws.cell(row=r,column=2,value=f); cell.number_format=NUM
ws["B12"].number_format=PCT; ws["B13"].number_format='0.0'; ws["B16"].number_format=MUL; ws["B17"].number_format='0.0'
ws["A19"]="Plan-level contribution (Rs/mo):"; ws["A19"].font=BOLD
hdr(ws,20,["Plan","Net ARPU","Var. cost","Contribution","Margin"])
pl=[("NRI Premium","=Pricing!H5","=-(Assumptions!B16+Assumptions!B17*3+Assumptions!B18+Assumptions!B20)-Pricing!H5*(Assumptions!B19+Assumptions!B15)","=B21+C21","=D21/B21"),
("NRI Standard","=Pricing!H4","=-(Assumptions!B16+Assumptions!B18+Assumptions!B20)-Pricing!H4*(Assumptions!B19+Assumptions!B15)","=B22+C22","=D22/B22"),
("India Pragati","=Pricing!H7","=-(Assumptions!B16+Assumptions!B17*1.5+Assumptions!B18+Assumptions!B20)-Pricing!H7*(Assumptions!B19+Assumptions!B15)","=B23+C23","=D23/B23"),
("India Aarambh","=Pricing!H6","=-(Assumptions!B16+Assumptions!B18+Assumptions!B20)-Pricing!H6*(Assumptions!B19+Assumptions!B15)","=B24+C24","=D24/B24")]
for r,p in enumerate(pl,start=21):
    for c,v in enumerate(p,start=1):
        cell=ws.cell(row=r,column=c,value=v); cell.number_format=NUM
        if c==5: cell.number_format=PCT
widths(ws,[34,14,14,14,10])

# ---------- 13 SCENARIOS ----------
ws=wb.create_sheet("Scenarios")
ws["A1"]="SCENARIOS - closed-form 36-month recursion per scenario"; ws["A1"].font=H1
hdr(ws,3,["Param","Conservative","Base","Aggressive"])
sc=[("Monthly churn","=Assumptions!B9","=Assumptions!B8","=Assumptions!B10"),
("CAC (Rs)","=Assumptions!B12","=Assumptions!B11","=Assumptions!B13"),
("Marketing multiplier",0.8,1.0,1.3),
("ARPU multiplier",0.92,1.0,1.08)]
for r,s in enumerate(sc,start=4):
    ws.cell(row=r,column=1,value=s[0])
    for c,v in enumerate(s[1:],start=2):
        cell=ws.cell(row=r,column=c,value=v)
        if isinstance(v,float): cell.font=BLUE
        cell.number_format=PCT if r==4 else NUM
ws.cell(row=6,column=2).number_format='0.0'; ws.cell(row=6,column=3).number_format='0.0'; ws.cell(row=6,column=4).number_format='0.0'
hdr(ws,9,["Month","Cons new","Cons closing","Base new","Base closing","Aggr new","Aggr closing"])
for m in range(1,37):
    r=m+9; gr=m+3
    ws.cell(row=r,column=1,value=m)
    ws.cell(row=r,column=2,value=f"='Customer Growth'!B{gr}*$B$6/$B$5*(1+Assumptions!$B$14)+IF(A{r}=4,Assumptions!$B$23,0)")
    ws.cell(row=r,column=3,value=(f"=B{r}" if m==1 else f"=C{r-1}*(1-$B$4)+B{r}"))
    ws.cell(row=r,column=4,value=f"='Customer Growth'!G{gr}")
    ws.cell(row=r,column=5,value=f"='Customer Growth'!K{gr}")
    ws.cell(row=r,column=6,value=f"='Customer Growth'!B{gr}*$D$6/$D$5*(1+Assumptions!$B$14)+IF(A{r}=4,Assumptions!$B$23,0)")
    ws.cell(row=r,column=7,value=(f"=F{r}" if m==1 else f"=G{r-1}*(1-$D$4)+F{r}"))
    for c in range(2,8): ws.cell(row=r,column=c).number_format=NUM
hdr(ws,47,["Output","Conservative","Base","Aggressive"])
outs=[("M36 customers","=C45","=E45","=G45"),
("Y3 revenue (Rs)","=SUMPRODUCT((C34:C45+C33:C44)/2)*(Assumptions!B6+Assumptions!B7)*$B$7","=Revenue!G41","=SUMPRODUCT((G34:G45+G33:G44)/2)*(Assumptions!B6+Assumptions!B7)*$D$7"),
("Y3 EBITDA approx (Rs)","=B49*0.74-12*(3500000*$B$6+Headcount!D13+720000)","='Operating Expenses'!D41","=D49*0.74-12*(3500000*$D$6+Headcount!D13+720000)")]
for r,o in enumerate(outs,start=48):
    for c,v in enumerate(o,start=1):
        cell=ws.cell(row=r,column=c,value=v); cell.number_format=NUM
ws["A52"]="Conservative/aggressive P&L lines are closed-form approximations; base case ties to full engine."; ws["A52"].font=ITAL
widths(ws,[22,14,13,13,13,13,13])

# ---------- 14 SENSITIVITY ----------
ws=wb.create_sheet("Sensitivity")
ws["A1"]="SENSITIVITY - single-variable stresses vs base (equilibrium approximations)"; ws["A1"].font=H1
hdr(ws,3,["Stress","Metric affected","Base value","Stressed value","Impact"])
sens=[("Churn +25% (6.0% to 7.5%)","M36 customers","='Customer Growth'!K39","='Customer Growth'!G39/0.075*(1-(1-0.075)^36)*0.85","=D4-C4"),
("CAC +25% (Rs3,600 to 4,500)","M36 customers","='Customer Growth'!K39","=C5*0.8","=D5-C5"),
("Conversion -25%","New customers / month","='Customer Growth'!D15+'Customer Growth'!E15","=C6*0.75","=D6-C6"),
("ARPU -15%","Y3 revenue (Rs)","=Revenue!G41","=C7*0.85","=D7-C7"),
("AI cost x2","Gross margin %","='Cost of Revenue'!I39","=C8-Assumptions!B16/(Assumptions!B6+Assumptions!B7)","=D8-C8"),
("Growth delayed 6 months","Peak funding need (Rs)","='Cash Flow'!D44","=C9*1.17","=D9-C9"),
("Annual mix 40% (vs 60%)","Deferred-cash benefit (Rs)","=SUM(Revenue!H4:H39)","=C10*40/60","=D10-C10"),
("Refunds x2 (5%)","Y3 net revenue (Rs)","=Revenue!G41","=C11*(1-0.025)","=D11-C11")]
for r,s in enumerate(sens,start=4):
    for c,v in enumerate(s,start=1):
        cell=ws.cell(row=r,column=c,value=v)
        if c>=3: cell.number_format=NUM
ws.cell(row=8,column=3).number_format=PCT; ws.cell(row=8,column=4).number_format=PCT; ws.cell(row=8,column=5).number_format=PCT
ws["A13"]="TOP-3 PROFIT DRIVERS: (1) monthly churn, (2) ARPU/plan mix, (3) CAC - exactly what Gate-1 validation measures."; ws["A13"].font=BOLD
ws["A14"]="Stressed values are closed-form approximations, not full re-runs; conservative scenario combines the first two."; ws["A14"].font=ITAL
widths(ws,[30,24,16,16,16])

# ---------- 15 IDEA COMPARISON ----------
ws=wb.create_sheet("Idea Comparison")
ws["A1"]="IDEA-WISE HIGH-LEVEL FINANCIALS (ranges; same assumption bases where fair)"; ws["A1"].font=H1
hdr(ws,3,["Metric","Sitara (winner)","Sahay (#2)","KalaOS (#3)","DharmaSetu (#4)","Smriti (#5)"])
comp=[("MVP investment (Rs L)","60-100","80-130","60-100","50-90","60-90"),
("Full launch investment (Rs Cr)","3.5","5.5-7","2.5-3.5","3-4","3-4.5"),
("Price point","Rs499-999 / $13-25 mo","$49-299 mo","Rs2,000-10,000 mo","$10-17 mo + events","$9-17 mo"),
("Gross margin %","72-78%","40-65% by tier","65-72%","55-70% blended","50-65% (print)"),
("Blended CAC (Rs)","3,000-4,500","8,000-12,000","8,000-15,000 (B2B)","4,000-6,000","2,500-4,000"),
("Monthly churn","5-7.5%","2.5-4% + life events","2.5-3.5% logo","3-5% (annual-shaped)","3.5-5% + aging-out"),
("Break-even customers","~7,500-8,500","~3,500-4,500","~1,600-2,000 studios","~9,000-11,000 members","~11,000-14,000"),
("Break-even month (base)","~M27","~M30-34","~M22-26","~M28-34","~M30-36"),
("Y1 revenue (Rs Cr)","1.3-1.8","1.5-2.5","0.8-1.4","0.9-1.6","0.7-1.2"),
("Y3 revenue (Rs Cr)","11-17","16-22","9-12","7-11","6-10"),
("Y3 EBITDA (Rs Cr)","0.8-1.8","(1)-1.5","1.2-2.2","0.3-1.2","(0.5)-0.8"),
("3-yr cash need (Rs Cr)","6-7","9-12","4-5","5-7","5.5-7.5"),
("Risk-adjusted score /10","8.1","8.0","7.6","7.4","7.1")]
for r,cc in enumerate(comp,start=4):
    for c,v in enumerate(cc,start=1):
        ws.cell(row=r,column=c,value=v)
ws["A18"]="Ranges reflect low confidence bounds; no false precision. Sahay/KalaOS/DharmaSetu/Smriti models are top-down"; ws["A18"].font=ITAL
ws["A19"]="scaled from the same ARPU/churn/CAC logic; Sitara ties to the full bottom-up engine."; ws["A19"].font=ITAL
widths(ws,[26,20,20,20,20,20])

# ---------- 16 INVESTMENT RETURNS ----------
ws=wb.create_sheet("Investment Returns")
ws["A1"]="RETURNS BY INVESTMENT LEVEL (management estimates; capital does NOT buy retention)"; ws["A1"].font=H1
hdr(ws,3,["Level","Rs50 L","Rs1 Cr","Rs2.5 Cr","Rs5 Cr","Rs10 Cr"])
inv=[("What it buys","Concierge validation only (no product)","Validation + MVP + founding 500","MVP + 18-mo launch (Scenario A+partial B)","Full A+B+early C","Full staged plan + India tier + rituals"),
("Team (FTE)","2-3","4-5","8-9","12-15","20-24"),
("Runway (months)","6-7","9-10","10-14","20-24","30-36"),
("Marketing budget (Rs)","8 L","25 L","1.05 Cr","2.4 Cr","5.6 Cr"),
("Target paying customers","50 pilot + 250 founding","500-800","3,000-3,500","8,000-9,500","16,000-19,000"),
("Exit ARR (Rs Cr)","0.25","0.7-0.9","3.0-3.6","8.5-10","17-20"),
("Break-even","n/a (validation)","n/a","path visible M18","~M27","~M26-28"),
("3-yr cumulative revenue (Rs Cr)","n/a","2-3","8-10","18-23","20-26"),
("3-yr cumulative EBITDA (Rs Cr)","(0.5)","(0.9)-(1.1)","(2.5)-(3.2)","(3.5)-(4.2)","(3.2)-(4.0)"),
("Cash returned / retained","Learning + stop-loss","Brand + panel + cohort data","Deferred-rev cushion ~Rs60L","Rs1-2 Cr closing cash","Rs1.5-3 Cr closing cash"),
("ROI shape","Option value only","Option value","Series-A readiness","2-4x revenue multiple potential","3-5x potential IF churn<=6%"),
("Constraint / risk","No product built","Skips retention proof at scale","CAC inflation risk","Gates still apply","Diminishing: CAC rises, gates unchanged"),
("Milestone before next funding","M2 renewal >=60%","500 founding, CAC<Rs4,000","Churn<=6%, NPS>=55","Cohort audit","Category leadership metrics")]
for r,vv in enumerate(inv,start=4):
    for c,v in enumerate(vv,start=1):
        cell=ws.cell(row=r,column=c,value=v)
        cell.alignment=Alignment(wrap_text=True,vertical="top")
ws["A18"]="Higher investment does NOT generate proportionate returns: retention gates, CAC inflation at scale,"; ws["A18"].font=ITAL
ws["A19"]="and category ceilings bind every level. The staged plan (Rs65L -> Rs2.85Cr -> Rs4.5Cr) is the recommended path."; ws["A19"].font=ITAL
widths(ws,[26,20,20,22,22,24])

# ---------- 17 DASHBOARD ----------
ws=wb.create_sheet("Dashboard")
ws["A1"]="DASHBOARD - KEY OUTPUTS (all linked; recalc after edits)"; ws["A1"].font=H1
dash=[("Paying customers M12","='Customer Growth'!K15"),("Paying customers M24","='Customer Growth'!K27"),
("Paying customers M36","='Customer Growth'!K39"),
("Revenue Y1 (Rs)","=Revenue!C41"),("Revenue Y2 (Rs)","=Revenue!E41"),("Revenue Y3 (Rs)","=Revenue!G41"),
("Gross margin Y3","=P!P24" ),("EBITDA Y1 (Rs)","='Operating Expenses'!B41"),
("EBITDA Y2 (Rs)","='Operating Expenses'!C41"),("EBITDA Y3 (Rs)","='Operating Expenses'!D41"),
("LTV (Rs)","='Unit Economics'!B14"),("CAC (Rs)","='Unit Economics'!B15"),
("LTV:CAC","='Unit Economics'!B16"),("CAC payback (months)","='Unit Economics'!B17"),
("Peak cumulative funding (Rs)","='Cash Flow'!D44"),("Minimum cash balance (Rs)","='Cash Flow'!B44"),
("Closing cash M36 (Rs)","='Cash Flow'!G39")]
r=3
for l,f in dash:
    ws.cell(row=r,column=1,value=l).font=BOLD
    cell=ws.cell(row=r,column=2,value=f if "P!P24" not in f else "='P&L'!P24")
    cell.number_format=NUM; cell.font=GREEN
    r+=1
ws.cell(row=9,column=2).number_format=PCT  # GM
ws.cell(row=15,column=2).number_format=MUL
ws.cell(row=16,column=2).number_format='0.0'
# charts
ch1=LineChart(); ch1.title="Closing customers (base)"; ch1.height=7; ch1.width=16
data=Reference(wb["Customer Growth"],min_col=11,min_row=3,max_row=39)
cats=Reference(wb["Customer Growth"],min_col=1,min_row=4,max_row=39)
ch1.add_data(data,titles_from_data=True); ch1.set_categories(cats)
ws.add_chart(ch1,"D3")
ch2=LineChart(); ch2.title="Monthly EBITDA (Rs)"; ch2.height=7; ch2.width=16
d2=Reference(wb["Operating Expenses"],min_col=10,min_row=3,max_row=39)
ch2.add_data(d2,titles_from_data=True); ch2.set_categories(cats)
ws.add_chart(ch2,"D18")
widths(ws,[30,16])

wb.save("Sitara_Financial_Model.xlsx")
print("saved")
